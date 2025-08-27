import streamlit as st
import pandas as pd
import os
import re
from io import BytesIO
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ==========================
# Utility Functions
# ==========================

def make_output_folders():
    """Ensure output folders exist."""
    os.makedirs("output/branchwise", exist_ok=True)
    os.makedirs("output/uniform", exist_ok=True)


def save_groups(groups_dict, group_type):
    """Save group CSVs into output/<group_type> folder."""
    folder = os.path.join("output", group_type)
    os.makedirs(folder, exist_ok=True)

    for gname, df in groups_dict.items():
        df.to_csv(os.path.join(folder, f"{gname}.csv"), index=False)

    return folder


def make_stats_table(groups_folder):
    """Generate department count statistics with 'Total' at the end."""
    group_files = sorted(
        [f for f in os.listdir(groups_folder) if f.endswith(".csv")],
        key=lambda x: int(re.search(r"\d+", x).group()) if re.search(r"\d+", x) else 9999
    )
    stats = {}
    for file in group_files:
        group_name = file.replace(".csv", "").upper()
        try:
            df = pd.read_csv(os.path.join(groups_folder, file))
        except Exception:
            continue
        if "Roll" not in df.columns:
            continue
        df["Department"] = df["Roll"].astype(str).str.extract(r"([A-Z]+)")
        counts = df["Department"].value_counts().to_dict()
        counts["Total"] = len(df)
        stats[group_name] = counts

    if stats:
        stats_df = pd.DataFrame(stats).T.fillna(0).astype(int)
        stats_df.index.name = "Group"

        # 🔥 Move "Total" to the end
        if "Total" in stats_df.columns:
            cols = [c for c in stats_df.columns if c != "Total"] + ["Total"]
            stats_df = stats_df[cols]
    else:
        stats_df = pd.DataFrame()

    return stats_df


def save_statistics(branchwise_folder, uniform_folder):
    """Save stats to Excel with 'Total' at the end and highlighted."""
    branchwise_stats = make_stats_table(branchwise_folder)
    uniform_stats = make_stats_table(uniform_folder)

    def reorder_total(df):
        if not df.empty and "Total" in df.columns:
            cols = [c for c in df.columns if c != "Total"] + ["Total"]
            return df[cols]
        return df

    branchwise_stats = reorder_total(branchwise_stats)
    uniform_stats = reorder_total(uniform_stats)

    output_file = os.path.join("output", "output.xlsx")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        start_row = 0
        if not branchwise_stats.empty:
            pd.DataFrame([["Branch Wise Mix"]]).to_excel(
                writer, sheet_name="Statistics", startrow=start_row, header=False, index=False
            )
            start_row += 2
            branchwise_stats.to_excel(writer, sheet_name="Statistics", startrow=start_row)
            start_row += len(branchwise_stats) + 3
        if not uniform_stats.empty:
            pd.DataFrame([["Uniform Mix"]]).to_excel(
                writer, sheet_name="Statistics", startrow=start_row, header=False, index=False
            )
            start_row += 2
            uniform_stats.to_excel(writer, sheet_name="Statistics", startrow=start_row)

    # 🔥 Highlight "Total" column
    wb = load_workbook(output_file)
    ws = wb["Statistics"]
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    bold_font = Font(bold=True)

    for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
        if col[0].value == "Total":  # column header
            for cell in col:
                cell.font = bold_font
                cell.fill = yellow_fill
    wb.save(output_file)

    return branchwise_stats, uniform_stats, output_file


def create_zip():
    """Create a ZIP of output folder for download."""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk("output"):
            for file in files:
                file_path = os.path.join(root, file)
                zf.write(file_path, os.path.relpath(file_path, "output"))
    zip_buffer.seek(0)
    return zip_buffer


# ==========================
# Streamlit App
# ==========================

st.set_page_config(page_title="Student Group Generator", layout="wide")

st.title("📊 Student Group Generator")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    st.write("### Preview of Uploaded Data")
    st.dataframe(df.head())

    if "Roll" not in df.columns:
        st.error("Excel must contain a 'Roll' column.")
    else:
        # Just for demo: split into groups of 5
        groups_branchwise = {
            f"Group_{i+1}": df.iloc[i:i+5] for i in range(0, len(df), 5)
        }
        groups_uniform = {
            f"Uniform_{i+1}": df.sample(min(5, len(df))) for i in range(len(groups_branchwise))
        }

        # Save groups
        make_output_folders()
        branchwise_folder = save_groups(groups_branchwise, "branchwise")
        uniform_folder = save_groups(groups_uniform, "uniform")

        # Save statistics
        branchwise_stats, uniform_stats, excel_file = save_statistics(branchwise_folder, uniform_folder)

        # Show tables
        st.subheader("📌 Branchwise Mix Statistics")
        if not branchwise_stats.empty:
            st.dataframe(branchwise_stats)
        else:
            st.info("No branchwise statistics available.")

        st.subheader("📌 Uniform Mix Statistics")
        if not uniform_stats.empty:
            st.dataframe(uniform_stats)
        else:
            st.info("No uniform statistics available.")

        # Download section
        st.subheader("⬇️ Download Results")
        zip_buffer = create_zip()
        st.download_button(
            label="Download All Outputs (ZIP)",
            data=zip_buffer,
            file_name="output.zip",
            mime="application/zip"
        )

        with open(excel_file, "rb") as f:
            st.download_button(
                label="Download Statistics Excel",
                data=f,
                file_name="output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
